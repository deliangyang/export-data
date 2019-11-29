from openpyxl import load_workbook
import xlrd
import openpyxl
import re
import json


wb = xlrd.open_workbook('590aa102d809444f.xlsx')
zl = wb.sheet_by_name('整理')

all_data = []
prev_row = [None for i in range(zl.ncols)]
for row_index in range(1, zl.nrows - 1):
    row = []
    for col_index in range(4):
        value = str(zl.cell(rowx=row_index, colx=col_index).value)
        if len(value) == 0:
            value = prev_row[col_index]
        row.append(value)
    prev_row = row
    all_data.append(row)

reg_detail = re.compile(r'^(.+?)(\d{11})([^$]+)')
reg_detail_2 = re.compile(r'^(\d{11})(\d{11})([^$]+)')


def split_item(detail):
    item = reg_detail_2.findall(detail)
    if len(item) <= 0:
        item = reg_detail.findall(detail)
    return item[0]


items = {}
for data in all_data:
    detail = data[1]
    name, mobile, address = split_item(detail)

    if detail not in items:
        tmp = {
            "name": name,
            "mobile": mobile,
            "address": address.strip(),
            "count": 0,
            "items": [],
        }
        items[detail] = tmp
    items[detail]["count"] += int(str(data[3]).replace('.0', ''))
    items[detail]["items"].append("%sx%s" % (data[2], data[3].replace('.0', '')))

wb = openpyxl.Workbook()
ws = wb.active


with open('xx.json', 'r') as f:
    address_order = json.loads(f.read().encode('utf-8'))
    f.close()

ws.append(['订单编号', '收件人', '手机', '地址', '发货信息', '数量'])
for item in items:
    t = items[item]
    order_no = ''
    name, mobile, address = t['name'], t['mobile'], t['address']
    if t['address'] in address_order:
        datum = address_order[t['address']]
        name_mobile = name + mobile
        if name_mobile in datum:
            order_no = ', '.join(datum[name_mobile])
        else:
            tmp = []
            for k in datum:
                tmp += datum[k]
            print(datum)
            order_no = ', '.join(tmp)
    ws.append([
        order_no, t['name'], t['mobile'], t['address'], ', '.join(t['items']), str(t['count']).replace('.0', '')
    ])

wb.save("result.xlsx")
